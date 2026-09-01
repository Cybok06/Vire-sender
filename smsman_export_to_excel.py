"""
SMS-MAN Export to Excel

Fetches SMS-MAN countries, services, and current prices using your API token, then creates one Excel file
with separate tabs:
  1. Countries
  2. Services
  3. Prices
  4. Raw Countries JSON
  5. Raw Services JSON
  6. Raw Prices JSON

Setup:
  pip install requests openpyxl python-dotenv

.env:
  SMSMAN_API_TOKEN=your_api_token_here

Run:
  python smsman_export_to_excel.py
"""

import json
import os
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://api.sms-man.com/control"
OUTPUT_FILE = "smsman_countries_services_prices.xlsx"
PRICE_REQUEST_DELAY_SECONDS = 0.35


def get_token() -> str:
    load_dotenv()
    token = os.getenv("SMSMAN_API_TOKEN", "").strip()
    if not token:
        raise RuntimeError("Missing SMSMAN_API_TOKEN in .env file")
    return token


def fetch_endpoint(endpoint: str, token: str, extra_params: Optional[Dict[str, Any]] = None) -> Any:
    url = f"{BASE_URL}/{endpoint}"
    params = {"token": token}
    if extra_params:
        params.update(extra_params)
    response = requests.get(url, params=params, timeout=30)
    response.raise_for_status()

    try:
        data = response.json()
    except ValueError as exc:
        raise RuntimeError(f"SMS-MAN returned non-JSON response for {endpoint}: {response.text[:300]}") from exc

    if isinstance(data, dict) and data.get("success") is False:
        raise RuntimeError(f"SMS-MAN error for {endpoint}: {data}")

    return data


def first_present(item: Dict[str, Any], keys: List[str], default: str = "") -> Any:
    for key in keys:
        value = item.get(key)
        if value not in (None, ""):
            return value
    return default


def normalize_rows(data: Any, preferred_columns: List[str]) -> List[Dict[str, Any]]:
    if isinstance(data, dict):
        data = [
            {**value, "id": value.get("id") or key} if isinstance(value, dict) else {"id": key, "value": value}
            for key, value in data.items()
        ]

    rows: List[Dict[str, Any]] = []
    if not isinstance(data, list):
        return rows

    for item in data:
        if isinstance(item, dict):
            row = {col: item.get(col, "") for col in preferred_columns}
            for key, value in item.items():
                if key not in row:
                    row[key] = value
            rows.append(row)
        else:
            rows.append({"value": item})

    return rows


def looks_like_price_item(value: Any) -> bool:
    return isinstance(value, dict) and ("cost" in value or "count" in value)


def iter_price_service_items(price_response: Dict[str, Any]):
    for group_or_service_id, grouped_value in price_response.items():
        if looks_like_price_item(grouped_value):
            yield group_or_service_id, grouped_value
            continue

        if not isinstance(grouped_value, dict):
            continue

        for service_id, price_item in grouped_value.items():
            yield service_id, price_item


def build_country_lookup(countries_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    lookup = {}
    for country in countries_rows:
        country_id = str(first_present(country, ["id", "country_id", "countryId"]))
        if country_id:
            lookup[country_id] = country
    return lookup


def build_service_lookup(services_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    lookup = {}
    for service in services_rows:
        service_id = str(first_present(service, ["id", "service_id", "application_id", "applicationId"]))
        if service_id:
            lookup[service_id] = service
    return lookup


def normalize_price_rows(
    prices_by_country: Dict[str, Any],
    countries_rows: List[Dict[str, Any]],
    services_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    country_lookup = build_country_lookup(countries_rows)
    service_lookup = build_service_lookup(services_rows)
    rows: List[Dict[str, Any]] = []

    for country_id, price_response in prices_by_country.items():
        if not isinstance(price_response, dict):
            continue

        country = country_lookup.get(str(country_id), {})
        country_title = first_present(country, ["title", "name", "country", "country_title"])

        for service_id, price_item in iter_price_service_items(price_response):
            if not isinstance(price_item, dict):
                price_item = {"value": price_item}

            service = service_lookup.get(str(service_id), {})
            rows.append({
                "provider": "smsman",
                "country_id": str(country_id),
                "country_title": country_title,
                "service_id": str(service_id),
                "service_title": first_present(service, ["title", "name", "service_title"]),
                "service_code": first_present(service, ["code", "service_code"]),
                "base_cost": price_item.get("cost", ""),
                "available_count": price_item.get("count", ""),
                "raw": price_item,
            })

    return rows


def fetch_prices_for_countries(
    token: str,
    countries_rows: List[Dict[str, Any]],
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    prices_by_country: Dict[str, Any] = {}
    errors: List[Dict[str, Any]] = []

    for index, country in enumerate(countries_rows, start=1):
        country_id = str(first_present(country, ["id", "country_id", "countryId"]))
        country_title = str(first_present(country, ["title", "name", "country", "country_title"], country_id))
        if not country_id:
            errors.append({
                "country_title": country_title,
                "error": "Country id missing from countries response.",
                "raw_country": country,
            })
            continue

        print(f"Fetching prices for {country_title} (country_id={country_id}) [{index}/{len(countries_rows)}]")
        try:
            prices_by_country[country_id] = fetch_endpoint("get-prices", token, {"country_id": country_id})
        except Exception as exc:
            errors.append({
                "country_id": country_id,
                "country_title": country_title,
                "error": str(exc),
            })
            print(f"  Failed for {country_title} (country_id={country_id}): {exc}")

        time.sleep(PRICE_REQUEST_DELAY_SECONDS)

    return prices_by_country, errors


def excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 3, 12), 42)


def write_table(ws, rows: List[Dict[str, Any]]):
    if not rows:
        ws.append(["No data returned"])
        return

    headers = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    ws.append(headers)
    for row in rows:
        ws.append([excel_value(row.get(header, "")) for header in headers])

    style_sheet(ws)


def write_raw_json(ws, title: str, data: Any):
    ws.append([title])
    ws.append([json.dumps(data, indent=2, ensure_ascii=False)])
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E40AF")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[2].height = 420
    ws.sheet_view.showGridLines = False


def create_excel(
    countries_data: Any,
    services_data: Any,
    prices_by_country: Dict[str, Any],
    price_errors: List[Dict[str, Any]],
    output_file: str,
):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    countries_rows = normalize_rows(countries_data, ["id", "title"])
    services_rows = normalize_rows(services_data, ["id", "name", "code"])
    price_rows = normalize_price_rows(prices_by_country, countries_rows, services_rows)

    ws_countries = wb.create_sheet("Countries")
    write_table(ws_countries, countries_rows)

    ws_services = wb.create_sheet("Services")
    write_table(ws_services, services_rows)

    ws_prices = wb.create_sheet("Prices")
    write_table(ws_prices, price_rows)

    ws_raw_countries = wb.create_sheet("Raw Countries JSON")
    write_raw_json(ws_raw_countries, "Countries raw API response", countries_data)

    ws_raw_services = wb.create_sheet("Raw Services JSON")
    write_raw_json(ws_raw_services, "Services raw API response", services_data)

    ws_raw_prices = wb.create_sheet("Raw Prices JSON")
    write_raw_json(ws_raw_prices, "Prices raw API responses and per-country errors", {
        "prices_by_country": prices_by_country,
        "errors": price_errors,
    })

    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["SMS-MAN Export Summary", ""])
    ws_summary.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Countries Count", len(countries_rows)])
    ws_summary.append(["Services Count", len(services_rows)])
    ws_summary.append(["Prices Count", len(price_rows)])
    ws_summary.append(["Countries with prices count", len(prices_by_country)])
    ws_summary.append(["Price Error Countries", len(price_errors)])
    ws_summary.append(["Countries Endpoint", f"{BASE_URL}/countries"])
    ws_summary.append(["Services Endpoint", f"{BASE_URL}/applications"])
    ws_summary.append(["Prices Endpoint", f"{BASE_URL}/get-prices?token=TOKEN&country_id=COUNTRY_ID"])
    ws_summary["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws_summary["A1"].fill = PatternFill("solid", fgColor="1E40AF")
    ws_summary.merge_cells("A1:B1")
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 60
    ws_summary.sheet_view.showGridLines = False

    wb.save(output_file)


def main():
    token = get_token()
    print("Fetching countries...")
    countries = fetch_endpoint("countries", token)
    print("Fetching services...")
    services = fetch_endpoint("applications", token)

    countries_rows = normalize_rows(countries, ["id", "title"])
    prices_by_country, price_errors = fetch_prices_for_countries(token, countries_rows)

    create_excel(countries, services, prices_by_country, price_errors, OUTPUT_FILE)
    print(f"Done: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
