"""
Import SMS-MAN country, service, and base price Excel data into MongoDB.

Run from the project root:
  python scripts/import_smsman_excel.py
"""

import json
import os
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient, UpdateOne
from pymongo.collection import Collection
from pymongo.errors import AutoReconnect, NetworkTimeout
from pymongo.server_api import ServerApi


PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = PROJECT_ROOT / "smsman_countries_services_prices.xlsx"
PROVIDER = "smsman"
BULK_BATCH_SIZE = 500
MAX_WRITE_RETRIES = 3


def load_environment() -> None:
    load_dotenv(PROJECT_ROOT / ".env")
    load_dotenv(PROJECT_ROOT / "backend" / ".env")


def get_mongo_client() -> MongoClient:
    mongo_uri = os.getenv("MONGO_URI", "").strip()
    if not mongo_uri or mongo_uri == "your_mongodb_uri":
        mongo_uri = os.getenv("MONGODB_URI", "").strip()

    if not mongo_uri:
        raise RuntimeError("Missing MONGO_URI or MONGODB_URI environment variable.")

    return MongoClient(mongo_uri, server_api=ServerApi("1"))


def get_database(client: MongoClient):
    db_name = os.getenv("MONGO_DB_NAME", "viresend").strip() or "viresend"
    return client.get_default_database(default=db_name)


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def parse_json_value(value: Any) -> Any:
    if value in (None, ""):
        return {}
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        try:
            return json.loads(value)
        except ValueError:
            return value
    return value


def to_string(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def to_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return 0


def read_sheet_rows(workbook_path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Missing required sheet: {sheet_name}")

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [to_string(value) for value in next(rows, [])]
    if not headers or headers == ["No data returned"]:
        return []

    records = []
    for values in rows:
        record = {
            header: values[index] if index < len(values) else ""
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)

    wb.close()
    return records


def raw_from_row(row: Dict[str, Any]) -> Dict[str, Any]:
    raw = parse_json_value(row.get("raw"))
    if isinstance(raw, dict):
        return raw
    return dict(row)


def country_document(row: Dict[str, Any], timestamp: datetime) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    country_id = to_string(row.get("id") or row.get("country_id") or row.get("provider_id"))
    title = to_string(row.get("title") or row.get("name") or row.get("country_title"))
    code = to_string(row.get("code") or row.get("iso") or row.get("short_name"))

    filter_doc = {
        "provider": PROVIDER,
        "$or": [
            {"country_id": country_id},
            {"provider_id": country_id},
        ],
    }
    update_doc = {
        "$set": {
            "provider": PROVIDER,
            "country_id": country_id,
            "title": title,
            "code": code,
            "is_active": True,
            "raw": raw_from_row(row),
            "updated_at": timestamp,
        },
        "$unset": {"provider_id": ""},
        "$setOnInsert": {"created_at": timestamp},
    }
    return filter_doc, update_doc


def service_document(row: Dict[str, Any], timestamp: datetime) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    provider_id = to_string(row.get("id") or row.get("provider_id") or row.get("service_id"))
    title = to_string(row.get("title") or row.get("name") or row.get("service_title"))
    name = to_string(row.get("name") or title)
    code = to_string(row.get("code") or row.get("service_code"))

    filter_doc = {"provider": PROVIDER, "provider_id": provider_id}
    update_doc = {
        "$set": {
            "provider": PROVIDER,
            "provider_id": provider_id,
            "title": title,
            "name": name,
            "code": code,
            "is_active": True,
            "raw": raw_from_row(row),
            "updated_at": timestamp,
        },
        "$setOnInsert": {"created_at": timestamp},
    }
    return filter_doc, update_doc


def price_document(row: Dict[str, Any], timestamp: datetime) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    country_id = to_string(row.get("country_id"))
    service_id = to_string(row.get("service_id"))

    filter_doc = {
        "provider": PROVIDER,
        "country_id": country_id,
        "service_id": service_id,
    }
    update_doc = {
        "$set": {
            "provider": PROVIDER,
            "country_id": country_id,
            "country_title": to_string(row.get("country_title")),
            "service_id": service_id,
            "service_title": to_string(row.get("service_title")),
            "service_code": to_string(row.get("service_code")),
            "base_cost": to_float(row.get("base_cost")),
            "available_count": to_int(row.get("available_count")),
            "currency": "GHS",
            "is_active": True,
            "raw": raw_from_row(row),
            "last_synced_at": timestamp,
            "updated_at": timestamp,
        },
        "$setOnInsert": {"created_at": timestamp},
    }
    return filter_doc, update_doc


def build_operations(
    rows: Iterable[Dict[str, Any]],
    document_builder,
    required_fields: Tuple[str, ...],
    timestamp: datetime,
) -> List[UpdateOne]:
    operations = []
    seen_keys = set()

    for row in rows:
        filter_doc, update_doc = document_builder(row, timestamp)
        set_doc = update_doc.get("$set", {})
        if any(not set_doc.get(field) for field in required_fields):
            continue

        key = tuple((field, set_doc.get(field)) for field in required_fields)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        operations.append(UpdateOne(filter_doc, update_doc, upsert=True))

    return operations


def chunked(items: List[UpdateOne], size: int) -> Iterable[List[UpdateOne]]:
    for index in range(0, len(items), size):
        yield items[index:index + size]


def write_batch(collection: Collection, batch: List[UpdateOne]):
    for attempt in range(1, MAX_WRITE_RETRIES + 1):
        try:
            return collection.bulk_write(batch, ordered=False)
        except (AutoReconnect, NetworkTimeout) as exc:
            if attempt == MAX_WRITE_RETRIES:
                raise
            wait_seconds = attempt * 2
            print(
                f"Temporary MongoDB connection issue while writing {collection.name}; "
                f"retrying in {wait_seconds}s ({attempt}/{MAX_WRITE_RETRIES}): {exc}"
            )
            time.sleep(wait_seconds)


def upsert_many(collection: Collection, operations: List[UpdateOne]) -> Dict[str, int]:
    if not operations:
        return {"matched": 0, "modified": 0, "upserted": 0, "total": 0}

    matched = 0
    modified = 0
    upserted = 0
    total_batches = (len(operations) + BULK_BATCH_SIZE - 1) // BULK_BATCH_SIZE

    for batch_number, batch in enumerate(chunked(operations, BULK_BATCH_SIZE), start=1):
        print(f"Writing {collection.name} batch {batch_number}/{total_batches} ({len(batch)} records)...")
        result = write_batch(collection, batch)
        matched += result.matched_count
        modified += result.modified_count
        upserted += len(result.upserted_ids)

    return {
        "matched": matched,
        "modified": modified,
        "upserted": upserted,
        "total": matched + upserted,
    }


def ensure_indexes(db) -> None:
    for index in db.smsman_countries.list_indexes():
        if index.get("name") == "provider_provider_id_unique":
            db.smsman_countries.drop_index("provider_provider_id_unique")

    db.smsman_countries.create_index(
        [("provider", 1), ("country_id", 1)],
        unique=True,
        name="provider_country_id_unique",
    )
    db.smsman_services.create_index(
        [("provider", 1), ("provider_id", 1)],
        unique=True,
        name="provider_provider_id_unique",
    )
    db.smsman_prices.create_index(
        [("provider", 1), ("country_id", 1), ("service_id", 1)],
        unique=True,
        name="provider_country_service_unique",
    )


def print_summary(label: str, stats: Dict[str, int]) -> None:
    print(
        f"{label}: {stats['total']} imported/updated "
        f"({stats['upserted']} inserted, {stats['matched']} updated or unchanged)"
    )


def main() -> None:
    load_environment()

    countries_rows = read_sheet_rows(EXCEL_FILE, "Countries")
    services_rows = read_sheet_rows(EXCEL_FILE, "Services")
    prices_rows = read_sheet_rows(EXCEL_FILE, "Prices")

    client = get_mongo_client()
    try:
        db = get_database(client)
        ensure_indexes(db)
        timestamp = now_utc()

        countries_ops = build_operations(
            countries_rows,
            country_document,
            ("provider", "country_id"),
            timestamp,
        )
        services_ops = build_operations(
            services_rows,
            service_document,
            ("provider", "provider_id"),
            timestamp,
        )
        prices_ops = build_operations(
            prices_rows,
            price_document,
            ("provider", "country_id", "service_id"),
            timestamp,
        )

        countries_stats = upsert_many(db.smsman_countries, countries_ops)
        services_stats = upsert_many(db.smsman_services, services_ops)
        prices_stats = upsert_many(db.smsman_prices, prices_ops)

        print_summary("Countries", countries_stats)
        print_summary("Services", services_stats)
        print_summary("Prices", prices_stats)
    finally:
        client.close()


if __name__ == "__main__":
    main()
