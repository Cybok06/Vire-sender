"""
Import SMS-MAN service image URLs into MongoDB.

Run from the project root:
  python scripts/import_smsman_service_images.py
"""

import csv
import os
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient, UpdateOne
from pymongo.collection import Collection
from pymongo.errors import AutoReconnect, NetworkTimeout
from pymongo.server_api import ServerApi


PROJECT_ROOT = Path(__file__).resolve().parent.parent
PROJECT_IMAGE_FILES = (
    PROJECT_ROOT / "smsman_service_image_urls.xlsx",
    PROJECT_ROOT / "smsman_services_images.xlsx",
    PROJECT_ROOT / "smsman_services_images.csv",
)
DOWNLOADS_IMAGE_FILES = (
    Path(r"C:\Users\cytec\Downloads\smsman_service_image_urls.xlsx"),
    Path(r"C:\Users\cytec\Downloads\smsman_services_images.xlsx"),
    Path(r"C:\Users\cytec\Downloads\smsman_services_images.csv"),
)
PROVIDER = "smsman"
BULK_BATCH_SIZE = 500
MAX_WRITE_RETRIES = 3
VERBOSE_SERVICE_LOGS = os.getenv("SMSMAN_IMAGE_IMPORT_VERBOSE", "true").strip().lower() in {"1", "true", "yes", "on"}


def load_environment() -> None:
    load_dotenv(PROJECT_ROOT / ".env")
    load_dotenv(PROJECT_ROOT / "backend" / ".env")


def get_mongo_client() -> MongoClient:
    mongo_uri = os.getenv("MONGO_URI", "").strip()
    if not mongo_uri or mongo_uri == "your_mongodb_uri":
        mongo_uri = os.getenv("MONGODB_URI", "").strip()
    if not mongo_uri:
        raise RuntimeError("Missing MONGO_URI or MONGODB_URI environment variable.")
    return MongoClient(mongo_uri, server_api=ServerApi("1"))


def get_database(client: MongoClient):
    db_name = os.getenv("MONGO_DB_NAME", "viresend").strip() or "viresend"
    return client.get_default_database(default=db_name)


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def clean(value: Any) -> str:
    return str(value or "").strip()


def normalize(value: Any) -> str:
    return clean(value).lower()


def image_file_path() -> Path:
    for path in (*PROJECT_IMAGE_FILES, *DOWNLOADS_IMAGE_FILES):
        if path.exists():
            return path
    return PROJECT_IMAGE_FILES[0]


def normalize_row_keys(row: Dict[str, Any]) -> Dict[str, str]:
    normalized = {clean(key).lower().replace(" ", "_"): clean(value) for key, value in row.items()}
    return {
        "no": normalized.get("no", ""),
        "service_name": normalized.get("service_name", "") or normalized.get("name", "") or normalized.get("service", ""),
        "smsman_service_code": normalized.get("smsman_service_code", "") or normalized.get("service_id", "") or normalized.get("provider_id", ""),
        "icon_code": normalized.get("icon_code", "") or normalized.get("code", ""),
        "image_url": normalized.get("image_url", "") or normalized.get("image", "") or normalized.get("url", ""),
    }


def read_csv_image_rows(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        return [
            normalize_row_keys(row)
            for row in csv.DictReader(csv_file)
            if clean(row.get("image_url") or row.get("image") or row.get("url"))
        ]


def read_excel_image_rows(path: Path) -> List[Dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [clean(value).lower().replace(" ", "_") for value in next(rows, [])]
        image_rows = []
        for values in rows:
            row = {
                header: values[index] if index < len(values) else ""
                for index, header in enumerate(headers)
                if header
            }
            normalized = normalize_row_keys(row)
            if normalized["image_url"]:
                image_rows.append(normalized)
        return image_rows
    finally:
        workbook.close()


def read_image_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"Image mapping file not found: {path}")

    if path.suffix.lower() == ".csv":
        return read_csv_image_rows(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_excel_image_rows(path)
    raise RuntimeError(f"Unsupported image mapping file type: {path.suffix}")


def find_default_image(rows: List[Dict[str, str]]) -> Optional[Dict[str, str]]:
    for row in rows:
        name = normalize(row.get("service_name"))
        if name in {"any other", "other", "other services", "default"}:
            return row
    for row in rows:
        if normalize(row.get("icon_code")) == "frame":
            return row
    return rows[0] if rows else None


def build_lookup(rows: List[Dict[str, str]]) -> Dict[str, Dict[str, Dict[str, str]]]:
    by_icon_code: Dict[str, Dict[str, str]] = {}
    by_provider_id: Dict[str, Dict[str, str]] = {}
    by_name: Dict[str, Dict[str, str]] = {}

    for row in rows:
        icon_code = normalize(row.get("icon_code"))
        provider_id = normalize(row.get("smsman_service_code"))
        service_name = normalize(row.get("service_name"))
        if icon_code and icon_code not in by_icon_code:
            by_icon_code[icon_code] = row
        if provider_id and provider_id not in by_provider_id:
            by_provider_id[provider_id] = row
        if service_name and service_name not in by_name:
            by_name[service_name] = row

    return {
        "by_icon_code": by_icon_code,
        "by_provider_id": by_provider_id,
        "by_name": by_name,
    }


def choose_image_row(service: Dict[str, Any], lookup: Dict[str, Dict[str, Dict[str, str]]], default_row: Dict[str, str]) -> Dict[str, str]:
    service_code = normalize(service.get("code"))
    provider_id = normalize(service.get("provider_id"))
    title = normalize(service.get("title"))
    name = normalize(service.get("name"))

    return (
        lookup["by_icon_code"].get(service_code)
        or lookup["by_provider_id"].get(provider_id)
        or lookup["by_name"].get(title)
        or lookup["by_name"].get(name)
        or default_row
    )


def chunked(items: List[UpdateOne], size: int) -> Iterable[List[UpdateOne]]:
    for index in range(0, len(items), size):
        yield items[index:index + size]


def write_batch(collection: Collection, batch: List[UpdateOne]):
    for attempt in range(1, MAX_WRITE_RETRIES + 1):
        try:
            return collection.bulk_write(batch, ordered=False)
        except (AutoReconnect, NetworkTimeout) as exc:
            if attempt == MAX_WRITE_RETRIES:
                raise
            wait_seconds = attempt * 2
            print(f"Temporary MongoDB issue while writing service images; retrying in {wait_seconds}s: {exc}")
            time.sleep(wait_seconds)


def main() -> None:
    load_environment()
    path = image_file_path()
    rows = read_image_rows(path)
    default_row = find_default_image(rows)
    if not default_row:
        raise RuntimeError("No image rows found in image mapping file.")

    lookup = build_lookup(rows)
    timestamp = now_utc()
    print(f"Loading service image mappings from: {path}")
    print(f"Image rows loaded: {len(rows)}")
    print(f"Default image: {default_row.get('image_url')} ({default_row.get('service_name') or 'default'})")

    client = get_mongo_client()
    try:
        db = get_database(client)
        services = list(db.smsman_services.find({"provider": PROVIDER}))
        print(f"Mongo services found: {len(services)}")
        operations: List[UpdateOne] = []
        matched_specific = 0

        for index, service in enumerate(services, start=1):
            image_row = choose_image_row(service, lookup, default_row)
            used_default = image_row is default_row
            if not used_default:
                matched_specific += 1
            service_title = clean(service.get("title") or service.get("name") or service.get("provider_id"))
            service_code = clean(service.get("code"))
            if VERBOSE_SERVICE_LOGS:
                match_type = "default" if used_default else "matched"
                print(
                    f"[{index}/{len(services)}] {match_type}: "
                    f"{service_title} code={service_code or '-'} "
                    f"provider_id={clean(service.get('provider_id')) or '-'} -> {image_row['image_url']}"
                )

            operations.append(UpdateOne(
                {"_id": service["_id"]},
                {"$set": {
                    "image": image_row["image_url"],
                    "image_url": image_row["image_url"],
                    "icon_code": image_row.get("icon_code", ""),
                    "image_source": path.name,
                    "updated_at": timestamp,
                }},
            ))

        modified = 0
        total_batches = (len(operations) + BULK_BATCH_SIZE - 1) // BULK_BATCH_SIZE
        for batch_number, batch in enumerate(chunked(operations, BULK_BATCH_SIZE), start=1):
            print(f"Writing service image batch {batch_number}/{total_batches} ({len(batch)} services)...")
            result = write_batch(db.smsman_services, batch)
            modified += result.modified_count
            print(
                f"Batch {batch_number}/{total_batches} complete: "
                f"matched={result.matched_count}, modified={result.modified_count}, total_modified={modified}"
            )

        print(f"Image mapping file: {path}")
        print(f"Services scanned: {len(services)}")
        print(f"Services with specific image match: {matched_specific}")
        print(f"Services using default image: {len(services) - matched_specific}")
        print(f"Services modified: {modified}")
    finally:
        client.close()


if __name__ == "__main__":
    main()
