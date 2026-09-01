"""Send the CYTECH voting SMS to the next unsent Excel recipients.

Preview the next batch (does not change the workbook):
    python send_cytech_bulk_sms.py

Send up to 100 unsent recipients and record progress in the workbook:
    python send_cytech_bulk_sms.py --send
"""

from __future__ import annotations

import argparse
import shutil
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("send message.xlsx")
SHEET_NAME = "Phone Numbers"
BATCH_SIZE = 100
SENDER_ID = "Nagonu"
MESSAGE = (
    "🎁 Claim your FREE 3GB Data by voting for CYTECH!\n"
    "Voting takes less than a minute and is FREE.\n"
    "✅ Click: https://startup.moolre.com/leaderboard/169\n"
    "✅ Verify your phone (OTP)\n"
    '✅ Vote for "CYTECH"'
)


def normalize_phone(value: object) -> str:
    return "".join(character for character in str(value or "") if character.isdigit())


def find_or_add_column(sheet, heading: str) -> int:
    for cell in sheet[1]:
        if str(cell.value or "").strip().casefold() == heading.casefold():
            return cell.column
    column = sheet.max_column + 1
    sheet.cell(1, column).value = heading
    return column


def get_unsent_rows(sheet, status_column: int) -> list[tuple[int, str]]:
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        phone = normalize_phone(sheet.cell(row_number, 1).value)
        status = str(sheet.cell(row_number, status_column).value or "").strip()
        if phone and not status:
            rows.append((row_number, phone))
            if len(rows) == BATCH_SIZE:
                break
    return rows


def get_arkesel_api_key() -> str:
    """Load the same Arkesel credentials used by the existing backend."""
    backend_path = Path(__file__).resolve().parent / "backend"
    sys.path.insert(0, str(backend_path))
    from app import create_app  # Imported only when actually sending.
    from services.sms_provider_service import normalize_sms_settings

    app = create_app()
    with app.app_context():
        settings = normalize_sms_settings(include_secret=True)

    if not settings.get("sms_enabled"):
        raise RuntimeError("SMS sending is disabled in the existing app settings.")
    if settings.get("active_sms_provider") != "arkesel":
        raise RuntimeError("The active SMS provider is not Arkesel.")
    if not settings.get("arkesel_enabled") or not settings.get("arkesel_api_key"):
        raise RuntimeError("Arkesel is not enabled or its API key is not configured.")
    return settings["arkesel_api_key"]


def main() -> int:
    parser = argparse.ArgumentParser(description="Send the next CYTECH SMS batch.")
    parser.add_argument("--send", action="store_true", help="Send SMS messages and update the workbook.")
    parser.add_argument("--delay", type=float, default=0.1, help="Seconds to wait between messages (default: 0.1).")
    args = parser.parse_args()

    if args.delay < 0:
        parser.error("--delay cannot be negative")
    if not WORKBOOK_PATH.is_file():
        print(f"Workbook not found: {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    workbook = load_workbook(WORKBOOK_PATH)
    if SHEET_NAME not in workbook.sheetnames:
        print(f"Sheet not found: {SHEET_NAME}", file=sys.stderr)
        return 1
    sheet = workbook[SHEET_NAME]
    status_column = find_or_add_column(sheet, "Status")
    sent_at_column = find_or_add_column(sheet, "Sent At")
    provider_message_column = find_or_add_column(sheet, "Provider Message")
    recipients = get_unsent_rows(sheet, status_column)

    if not recipients:
        print("No unsent phone numbers remain.")
        return 0

    print(f"Next batch: {len(recipients)}/{BATCH_SIZE} unsent numbers. Sender ID: {SENDER_ID}")
    if not args.send:
        print("Preview only: no SMS was sent and the workbook was not changed. Use --send to start.")
        return 0

    backup_path = WORKBOOK_PATH.with_name(f"{WORKBOOK_PATH.stem}.before-send.xlsx")
    if not backup_path.exists():
        shutil.copy2(WORKBOOK_PATH, backup_path)
        print(f"Backup created: {backup_path.name}")

    api_key = get_arkesel_api_key()
    from services.arkesel_service import send_sms

    successes = failures = 0
    total = len(recipients)
    for position, (row_number, phone) in enumerate(recipients, start=1):
        result = send_sms(api_key, SENDER_ID, MESSAGE, [phone])
        provider_message = result.get("message", "")
        if result.get("success"):
            sheet.cell(row_number, status_column).value = "SENT"
            sheet.cell(row_number, sent_at_column).value = datetime.now(timezone.utc).isoformat()
            sheet.cell(row_number, provider_message_column).value = provider_message
            workbook.save(WORKBOOK_PATH)
            successes += 1
            print(f"[{position}/{total}] SENT {phone}")
        else:
            # A failure is recorded too, preventing accidental repeated charges. Clear this
            # status manually only if you deliberately want to retry that recipient.
            sheet.cell(row_number, status_column).value = "FAILED"
            sheet.cell(row_number, provider_message_column).value = provider_message
            workbook.save(WORKBOOK_PATH)
            failures += 1
            print(f"[{position}/{total}] FAILED {phone}: {provider_message}")
        if position < total and args.delay:
            time.sleep(args.delay)

    print(f"Batch complete: {successes} sent, {failures} failed, {total} processed.")
    return 0 if failures == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
