"""Report duplicate phone numbers from an Excel workbook.

Usage:
    python scan_duplicate_numbers.py "send message.xlsx"
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def normalize_phone(value: object) -> str:
    """Compare phone numbers using digits only."""
    return re.sub(r"\D", "", str(value or ""))


def main() -> int:
    workbook_path = Path(sys.argv[1] if len(sys.argv) > 1 else "send message.xlsx")
    if not workbook_path.is_file():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    occurrences: dict[str, list[int]] = defaultdict(list)

    for row_number, (value,) in enumerate(
        sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2
    ):
        number = normalize_phone(value)
        if number:
            occurrences[number].append(row_number)

    duplicates = {number: rows for number, rows in occurrences.items() if len(rows) > 1}
    total = sum(len(rows) for rows in occurrences.values())
    print(f"Scanned {total:,} phone numbers in {workbook_path.name}.")
    print(f"Unique phone numbers: {len(occurrences):,}")
    print(f"Duplicate phone numbers: {len(duplicates):,}")

    for number, rows in duplicates.items():
        print(f"{number}: rows {', '.join(map(str, rows))}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
