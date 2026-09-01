import csv
import io
import re
import secrets
from decimal import Decimal, InvalidOperation

from bson import ObjectId
from openpyxl import load_workbook
from pymongo import ReturnDocument
from pymongo.errors import BulkWriteError, DuplicateKeyError
from flask import Blueprint, current_app, jsonify, request

from utils.auth import require_admin, require_auth, users_collection
from utils.notifications import create_notification
from utils.security import clean_string, is_valid_email, now_utc
from utils.service_control import check_service_available
from services.cloudflare_images_service import CloudflareImageError, upload_image

admin_contact_packages_bp = Blueprint("admin_contact_packages", __name__, url_prefix="/api/admin/contact-packages")
admin_contact_marketplace_bp = Blueprint("admin_contact_marketplace", __name__, url_prefix="/api/admin/contact-marketplace")
contact_marketplace_bp = Blueprint("contact_marketplace", __name__, url_prefix="/api/contact-marketplace")

MAX_UPLOAD_BYTES = 2 * 1024 * 1024
MAX_UPLOAD_ROWS = 10000
MAX_COVER_IMAGE_BYTES = 5 * 1024 * 1024
ALLOWED_COVER_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}


def packages_collection():
    return current_app.config["DB"].contact_packages


def marketplace_contacts_collection():
    return current_app.config["DB"].marketplace_contacts


def purchases_collection():
    return current_app.config["DB"].contact_package_purchases


def contacts_collection():
    return current_app.config["DB"].contacts


def contact_import_jobs_collection():
    return current_app.config["DB"].contact_import_jobs


def wallet_transactions_collection():
    return current_app.config["DB"].wallet_transactions


def iso(value):
    return value.isoformat() if value else None


def money(value, fallback=0.0):
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return fallback
    return round(float(amount), 2)


def create_reference(prefix="CP"):
    return f"{prefix}-{now_utc().strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(3).upper()}"


def normalize_phone(value: str) -> str | None:
    phone = re.sub(r"[^\d+]", "", value or "")
    if phone.startswith("+"):
        phone = phone[1:]
    if phone.startswith("0") and len(phone) == 10:
        phone = "233" + phone[1:]
    if phone.startswith("2330"):
        phone = "233" + phone[4:]
    if not phone.startswith("233") or not phone.isdigit() or len(phone) != 12:
        return None
    return phone


def mask_phone(value: str) -> str:
    phone = normalize_phone(value) or re.sub(r"\D", "", value or "")
    if phone.startswith("233") and len(phone) == 12:
        phone = "0" + phone[3:]
    if len(phone) < 4:
        return phone
    return f"{phone[:2]}******{phone[-2:]}"


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-")
    return slug or f"package-{secrets.token_hex(3)}"


def get_current_user(payload):
    user_id = payload.get("user_id") or payload.get("sub")
    try:
        object_id = ObjectId(user_id)
    except Exception:
        return None
    return users_collection().find_one({"_id": object_id})


def require_active_user(payload):
    user = get_current_user(payload)
    if not user:
        return None, ({"success": False, "message": "User account not found."}, 404)
    if user.get("account_status") != "active":
        return None, ({"success": False, "message": "Your account is not active."}, 403)
    return user, None


def find_package(package_id: str):
    query = {"package_id": clean_string(package_id)}
    try:
        query = {"$or": [query, {"_id": ObjectId(package_id)}]}
    except Exception:
        pass
    return packages_collection().find_one(query)


def package_preview(package_id: str, limit=3):
    contacts = marketplace_contacts_collection().find({"package_id": package_id}).sort("created_at", 1).limit(limit)
    return [
        {
            "name": item.get("name", ""),
            "phone": mask_phone(item.get("normalized_phone") or item.get("phone", "")),
            "location": item.get("location", ""),
        }
        for item in contacts
    ]


def safe_package(package: dict, purchased=False, include_admin=False) -> dict:
    payload = {
        "id": package.get("package_id") or str(package.get("_id")),
        "package_id": package.get("package_id") or str(package.get("_id")),
        "title": package.get("title", ""),
        "slug": package.get("slug", ""),
        "description": package.get("description", ""),
        "cover_image_url": package.get("cover_image_url", ""),
        "category": package.get("category", ""),
        "price": money(package.get("price")),
        "currency": package.get("currency", "GHS"),
        "total_contacts": int(package.get("total_contacts", 0) or 0),
        "sample_contacts_preview": package_preview(package.get("package_id", "")),
        "status": package.get("status", "inactive"),
        "purchased": purchased,
        "created_at": iso(package.get("created_at")),
        "updated_at": iso(package.get("updated_at")),
    }
    if include_admin:
        payload["created_by_admin_id"] = package.get("created_by_admin_id", "admin")
    return payload


def safe_purchase(purchase: dict, user: dict | None = None) -> dict:
    return {
        "id": purchase.get("purchase_id") or str(purchase.get("_id")),
        "purchase_id": purchase.get("purchase_id") or str(purchase.get("_id")),
        "user_id": str(purchase.get("user_id")) if purchase.get("user_id") else None,
        "user_name": user.get("full_name", "") if user else purchase.get("user_name", ""),
        "user_email": user.get("email") if user else purchase.get("user_email", ""),
        "package_id": purchase.get("package_id", ""),
        "package_title": purchase.get("package_title", ""),
        "price": money(purchase.get("price")),
        "currency": purchase.get("currency", "GHS"),
        "total_contacts": int(purchase.get("total_contacts", 0) or 0),
        "wallet_before": money(purchase.get("wallet_before")),
        "wallet_after": money(purchase.get("wallet_after")),
        "status": purchase.get("status", "completed"),
        "import_summary": purchase.get("import_summary") or {},
        "created_at": iso(purchase.get("created_at")),
    }


def clean_package_payload(data: dict, partial=False):
    title = clean_string(data.get("title", ""))
    description = clean_string(data.get("description", ""))
    category = clean_string(data.get("category", ""))
    status = clean_string(data.get("status", "inactive")).lower()
    price = money(data.get("price"), -1)
    errors = {}

    if not partial or "title" in data:
        if not title:
            errors["title"] = "Package name is required."
    if not partial or "category" in data:
        if not category:
            errors["category"] = "Category is required."
    if not partial or "price" in data:
        if price < 0:
            errors["price"] = "Enter a valid price."
    if status not in {"active", "inactive"}:
        errors["status"] = "Status must be active or inactive."

    return errors, {
        "title": title,
        "description": description,
        "category": category,
        "price": price,
        "currency": "GHS",
        "status": status,
    }


def parse_contact_row(row: dict):
    name = clean_string(row.get("name", ""))[:120]
    phone = clean_string(row.get("phone", ""))
    normalized_phone = normalize_phone(phone)
    email = clean_string(row.get("email", "")).lower()
    location = clean_string(row.get("location", ""))[:120]
    notes = clean_string(row.get("notes", ""))[:500]

    if not normalized_phone:
        return None, "A valid Ghana phone number is required."
    if email and not is_valid_email(email):
        return None, "Email is invalid."

    return {
        "name": name,
        "phone": normalized_phone,
        "normalized_phone": normalized_phone,
        "email": email,
        "location": location,
        "notes": notes,
    }, None


def normalize_upload_headers(row: dict) -> dict:
    aliases = {
        "name": "name",
        "contact name": "name",
        "phone": "phone",
        "phone number": "phone",
        "phone no": "phone",
        "mobile": "phone",
        "mobile number": "phone",
        "email": "email",
        "location": "location",
        "notes": "notes",
    }
    normalized = {}
    for key, value in row.items():
        header = re.sub(r"\s+", " ", str(key or "").strip().lower())
        field = aliases.get(header)
        if field:
            normalized[field] = value

    phone = normalized.get("phone")
    if isinstance(phone, float) and phone.is_integer():
        phone = int(phone)
    if isinstance(phone, int):
        phone = str(phone)
        if len(phone) == 9:
            phone = "0" + phone
        normalized["phone"] = phone
    return normalized


def read_excel_contacts(content: bytes):
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        normalized_headers = {re.sub(r"\s+", " ", str(header or "").strip().lower()) for header in headers}
        if "phone number" not in normalized_headers:
            raise ValueError('Excel file must contain a "Phone number" column.')
        contacts = []
        for values in rows:
            row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
            contacts.append(normalize_upload_headers(row))
        return contacts
    finally:
        workbook.close()


def refresh_package_counts(package_id: str):
    total = marketplace_contacts_collection().count_documents({"package_id": package_id})
    preview = package_preview(package_id)
    packages_collection().update_one(
        {"package_id": package_id},
        {"$set": {"total_contacts": total, "sample_contacts_preview": preview, "updated_at": now_utc()}},
    )
    return total


@admin_contact_packages_bp.post("")
@require_admin
def create_package(payload):
    data = request.get_json(silent=True) or {}
    errors, cleaned = clean_package_payload(data)
    if errors:
        return {"success": False, "message": "Please correct the highlighted fields.", "errors": errors}, 400

    now = now_utc()
    base_slug = slugify(cleaned["title"])
    slug = base_slug
    suffix = 1
    while packages_collection().find_one({"slug": slug}):
        suffix += 1
        slug = f"{base_slug}-{suffix}"

    package = {
        "package_id": create_reference("PKG"),
        "slug": slug,
        **cleaned,
        "total_contacts": 0,
        "sample_contacts_preview": [],
        "created_by_admin_id": payload.get("user_id", "admin"),
        "created_at": now,
        "updated_at": now,
    }
    packages_collection().insert_one(package)
    return jsonify({"success": True, "message": "Contact package created.", "package": safe_package(package, include_admin=True)}), 201


@admin_contact_packages_bp.get("")
@require_admin
def list_admin_packages(payload):
    packages = packages_collection().find({}).sort("created_at", -1)
    return jsonify({"success": True, "packages": [safe_package(package, include_admin=True) for package in packages]})


@admin_contact_packages_bp.get("/<package_id>")
@require_admin
def get_admin_package(payload, package_id):
    package = find_package(package_id)
    if not package:
        return {"success": False, "message": "Contact package not found."}, 404
    return jsonify({"success": True, "package": safe_package(package, include_admin=True)})


@admin_contact_packages_bp.put("/<package_id>")
@require_admin
def update_package(payload, package_id):
    package = find_package(package_id)
    if not package:
        return {"success": False, "message": "Contact package not found."}, 404

    data = request.get_json(silent=True) or {}
    errors, cleaned = clean_package_payload(data, partial=True)
    if errors:
        return {"success": False, "message": "Please correct the highlighted fields.", "errors": errors}, 400

    update = {key: value for key, value in cleaned.items() if key in data or key == "currency"}
    if "title" in update and update["title"] != package.get("title"):
        update["slug"] = slugify(update["title"])
    update["updated_at"] = now_utc()
    packages_collection().update_one({"_id": package["_id"]}, {"$set": update})
    updated = packages_collection().find_one({"_id": package["_id"]})
    return jsonify({"success": True, "message": "Contact package updated.", "package": safe_package(updated, include_admin=True)})


@admin_contact_packages_bp.delete("/<package_id>")
@require_admin
def delete_package(payload, package_id):
    package = find_package(package_id)
    if not package:
        return {"success": False, "message": "Contact package not found."}, 404

    action = clean_string(request.args.get("action", "deactivate")).lower()
    if action == "delete":
        data = request.get_json(silent=True) or {}
        confirmation = clean_string(data.get("confirmation", ""))
        expected_confirmation = f"Delete {package.get('title', '')}"
        if confirmation != expected_confirmation:
            return {
                "success": False,
                "message": f'Type "{expected_confirmation}" to permanently delete this package.',
            }, 400
        marketplace_contacts_collection().delete_many({"package_id": package["package_id"]})
        packages_collection().delete_one({"_id": package["_id"]})
        return jsonify({"success": True, "message": "Contact package deleted."})

    packages_collection().update_one({"_id": package["_id"]}, {"$set": {"status": "inactive", "updated_at": now_utc()}})
    return jsonify({"success": True, "message": "Contact package deactivated."})


@admin_contact_packages_bp.post("/<package_id>/upload")
@require_admin
def upload_package_contacts(payload, package_id):
    package = find_package(package_id)
    if not package:
        return {"success": False, "message": "Contact package not found."}, 404

    rows = []
    job_id = clean_string(request.form.get("job_id", "")) if request.files.get("file") else ""
    if request.files.get("file"):
        file = request.files["file"]
        content = file.read(MAX_UPLOAD_BYTES + 1)
        if len(content) > MAX_UPLOAD_BYTES:
            return {"success": False, "message": "Upload is too large. Maximum size is 2 MB."}, 400
        filename = (file.filename or "").lower()
        try:
            if filename.endswith(".xlsx"):
                rows = read_excel_contacts(content)
            elif filename.endswith(".csv"):
                text = content.decode("utf-8-sig")
                rows = [normalize_upload_headers(row) for row in csv.DictReader(io.StringIO(text))]
            else:
                return {"success": False, "message": "Upload a CSV or Excel (.xlsx) file."}, 400
        except ValueError as exc:
            return {"success": False, "message": str(exc) or "The uploaded file could not be read."}, 400
        except (UnicodeDecodeError, KeyError, OSError):
            return {"success": False, "message": "The uploaded file could not be read."}, 400
        except Exception:
            return {"success": False, "message": "The uploaded file could not be read."}, 400
    else:
        data = request.get_json(silent=True) or {}
        rows = data.get("contacts") or []

    if not rows:
        return {"success": False, "message": "No contacts were provided."}, 400
    requested_limit = 0
    if request.files.get("file"):
        raw_limit = clean_string(request.form.get("max_contacts", ""))
        if raw_limit:
            try:
                requested_limit = int(raw_limit)
            except ValueError:
                return {"success": False, "message": "Enter a valid maximum contact count."}, 400
            if requested_limit < 1 or requested_limit > MAX_UPLOAD_ROWS:
                return {"success": False, "message": f"Maximum contacts must be between 1 and {MAX_UPLOAD_ROWS}."}, 400
            rows = rows[:requested_limit]
    if len(rows) > MAX_UPLOAD_ROWS:
        return {"success": False, "message": f"Upload is limited to {MAX_UPLOAD_ROWS} rows. Enter a smaller maximum contact count."}, 400

    now = now_utc()
    if job_id:
        contact_import_jobs_collection().update_one(
            {"job_id": job_id},
            {"$set": {"job_id": job_id, "admin_id": payload.get("user_id"), "package_id": package["package_id"], "status": "processing", "processed": 0, "total": len(rows), "imported": 0, "updated_at": now}, "$setOnInsert": {"created_at": now}},
            upsert=True,
        )
    imported = 0
    duplicate_skipped = 0
    failed = 0
    seen = set()

    documents = []
    for row in rows:
        cleaned, error = parse_contact_row(row)
        if error:
            failed += 1
            continue
        if cleaned["normalized_phone"] in seen:
            duplicate_skipped += 1
            continue
        seen.add(cleaned["normalized_phone"])
        documents.append({"package_id": package["package_id"], **cleaned, "created_at": now})

    batch_size = 500
    attempted_so_far = 0
    for start in range(0, len(documents), batch_size):
        batch = documents[start:start + batch_size]
        try:
            result = marketplace_contacts_collection().insert_many(batch, ordered=False)
            inserted = len(result.inserted_ids)
            imported += inserted
        except BulkWriteError as exc:
            inserted = int(exc.details.get("nInserted", 0) or 0)
            imported += inserted
            duplicate_skipped += len(exc.details.get("writeErrors") or [])
        except Exception:
            failed += len(batch)
        attempted_so_far += len(batch)
        if job_id:
            contact_import_jobs_collection().update_one({"job_id": job_id}, {"$set": {"processed": min(len(rows), attempted_so_far), "imported": imported, "duplicate_skipped": duplicate_skipped, "failed": failed, "updated_at": now_utc()}})

    total = refresh_package_counts(package["package_id"])
    if job_id:
        contact_import_jobs_collection().update_one({"job_id": job_id}, {"$set": {"status": "completed", "processed": len(rows), "total": len(rows), "imported": imported, "duplicate_skipped": duplicate_skipped, "failed": failed, "updated_at": now_utc()}})
    return jsonify({
        "success": True,
        "message": f"Uploaded {imported} contacts. {duplicate_skipped} duplicates skipped.",
        "summary": {
            "total_contacts": total,
            "imported_contacts": imported,
            "duplicate_skipped": duplicate_skipped,
            "failed_contacts": failed,
            "requested_limit": requested_limit or None,
        },
    })


@admin_contact_packages_bp.get("/upload-status/<job_id>")
@require_admin
def package_upload_status(payload, job_id):
    job = contact_import_jobs_collection().find_one({"job_id": clean_string(job_id), "admin_id": payload.get("user_id")})
    if not job:
        return jsonify({"success": True, "status": "waiting", "processed": 0, "total": 0, "imported": 0})
    return jsonify({"success": True, "status": job.get("status", "processing"), "processed": int(job.get("processed", 0) or 0), "total": int(job.get("total", 0) or 0), "imported": int(job.get("imported", 0) or 0), "duplicate_skipped": int(job.get("duplicate_skipped", 0) or 0), "failed": int(job.get("failed", 0) or 0)})


@admin_contact_packages_bp.post("/<package_id>/cover-image")
@require_admin
def upload_package_cover_image(payload, package_id):
    package = find_package(package_id)
    if not package:
        return {"success": False, "message": "Contact package not found."}, 404
    file = request.files.get("file")
    if not file or not file.filename:
        return {"success": False, "message": "Select a cover image to upload."}, 400
    if file.mimetype not in ALLOWED_COVER_IMAGE_TYPES:
        return {"success": False, "message": "Upload a JPG, PNG, WebP, or GIF image."}, 400
    content = file.read(MAX_COVER_IMAGE_BYTES + 1)
    if len(content) > MAX_COVER_IMAGE_BYTES:
        return {"success": False, "message": "Cover image must be 5 MB or smaller."}, 400
    try:
        uploaded = upload_image(content, file.filename, file.mimetype)
    except CloudflareImageError as exc:
        return {"success": False, "message": str(exc)}, 502
    packages_collection().update_one(
        {"_id": package["_id"]},
        {"$set": {"cover_image_id": uploaded["id"], "cover_image_url": uploaded["url"], "updated_at": now_utc()}},
    )
    updated = packages_collection().find_one({"_id": package["_id"]})
    return jsonify({"success": True, "message": "Cover image uploaded.", "package": safe_package(updated, include_admin=True)})


@admin_contact_packages_bp.get("/<package_id>/contacts")
@require_admin
def list_package_contacts(payload, package_id):
    package = find_package(package_id)
    if not package:
        return {"success": False, "message": "Contact package not found."}, 404
    contacts = marketplace_contacts_collection().find({"package_id": package["package_id"]}).sort("created_at", -1).limit(1000)
    return jsonify({"success": True, "contacts": [
        {
            "id": str(item["_id"]),
            "name": item.get("name", ""),
            "phone": item.get("phone", ""),
            "normalized_phone": item.get("normalized_phone", ""),
            "email": item.get("email", ""),
            "location": item.get("location", ""),
            "notes": item.get("notes", ""),
            "created_at": iso(item.get("created_at")),
        }
        for item in contacts
    ]})


@admin_contact_marketplace_bp.get("/stats")
@require_admin
def admin_marketplace_stats(payload):
    packages = list(packages_collection().find({}))
    purchases = list(purchases_collection().find({"status": "completed"}))
    total_revenue = sum(money(item.get("price")) for item in purchases)
    top = []
    for package in packages:
        package_purchases = [p for p in purchases if p.get("package_id") == package.get("package_id")]
        top.append({
            "package_id": package.get("package_id"),
            "title": package.get("title", ""),
            "purchases": len(package_purchases),
            "revenue": round(sum(money(p.get("price")) for p in package_purchases), 2),
        })
    top.sort(key=lambda item: (item["purchases"], item["revenue"]), reverse=True)
    recent = list(purchases_collection().find({}).sort("created_at", -1).limit(8))
    return jsonify({"success": True, "stats": {
        "total_packages": len(packages),
        "active_packages": sum(1 for package in packages if package.get("status") == "active"),
        "total_contacts_uploaded": sum(int(package.get("total_contacts", 0) or 0) for package in packages),
        "total_purchases": len(purchases),
        "total_revenue": round(total_revenue, 2),
        "top_selling_packages": top[:5],
        "recent_purchases": [safe_purchase(purchase) for purchase in recent],
    }})


@admin_contact_marketplace_bp.get("/purchases")
@require_admin
def admin_marketplace_purchases(payload):
    purchases = list(purchases_collection().find({}).sort("created_at", -1).limit(500))
    user_ids = [p.get("user_id") for p in purchases if isinstance(p.get("user_id"), ObjectId)]
    users = {
        user["_id"]: user
        for user in users_collection().find({"_id": {"$in": user_ids}}, {"password_hash": 0})
    } if user_ids else {}
    return jsonify({"success": True, "purchases": [safe_purchase(purchase, users.get(purchase.get("user_id"))) for purchase in purchases]})


@contact_marketplace_bp.get("/packages")
@require_auth
def list_user_packages(payload):
    user, error = require_active_user(payload)
    if error:
        return error

    query = {"status": "active"}
    search = clean_string(request.args.get("search", ""))
    category = clean_string(request.args.get("category", ""))
    if search:
        query["$or"] = [
            {"title": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}},
        ]
    if category:
        query["category"] = category

    purchases = purchases_collection().find({"user_id": user["_id"], "status": "completed"}, {"package_id": 1})
    purchased_ids = {purchase.get("package_id") for purchase in purchases}
    packages = packages_collection().find(query).sort("created_at", -1)
    categories = sorted(packages_collection().distinct("category", {"status": "active"}))
    return jsonify({
        "success": True,
        "packages": [safe_package(package, purchased=package.get("package_id") in purchased_ids) for package in packages],
        "categories": categories,
    })


@contact_marketplace_bp.get("/packages/<package_id>")
@require_auth
def get_user_package(payload, package_id):
    user, error = require_active_user(payload)
    if error:
        return error
    package = find_package(package_id)
    if not package or package.get("status") != "active":
        return {"success": False, "message": "Contact package not found."}, 404
    purchased = purchases_collection().find_one({"user_id": user["_id"], "package_id": package["package_id"], "status": "completed"})
    return jsonify({"success": True, "package": safe_package(package, purchased=bool(purchased))})


@contact_marketplace_bp.get("/purchases")
@require_auth
def user_purchases(payload):
    user, error = require_active_user(payload)
    if error:
        return error
    purchases = purchases_collection().find({"user_id": user["_id"]}).sort("created_at", -1)
    return jsonify({"success": True, "purchases": [safe_purchase(purchase) for purchase in purchases]})


@contact_marketplace_bp.post("/packages/<package_id>/buy")
@require_auth
def buy_package(payload, package_id):
    locked = check_service_available("buy_contacts")
    if locked:
        return locked
    user, error = require_active_user(payload)
    if error:
        return error

    package = find_package(package_id)
    if not package or package.get("status") != "active":
        return {"success": False, "message": "Contact package is not available."}, 404
    if int(package.get("total_contacts", 0) or 0) <= 0:
        return {"success": False, "message": "This contact package has no contacts yet."}, 400

    existing = purchases_collection().find_one({"user_id": user["_id"], "package_id": package["package_id"], "status": "completed"})
    if existing:
        return {"success": False, "message": "You have already purchased this contact package.", "purchase": safe_purchase(existing)}, 409

    price = money(package.get("price"))
    wallet_before = money(user.get("wallet_balance"))
    if wallet_before < price:
        return {"success": False, "message": "Insufficient wallet balance.", "balance": wallet_before}, 400

    wallet_after = round(wallet_before - price, 2)
    now = now_utc()
    reference = create_reference("CP")
    updated_user = users_collection().find_one_and_update(
        {"_id": user["_id"], "wallet_balance": {"$gte": price}},
        {"$inc": {"wallet_balance": -price}, "$set": {"updated_at": now}},
        return_document=ReturnDocument.AFTER,
    )
    if not updated_user:
        return {"success": False, "message": "Insufficient wallet balance."}, 400
    wallet_after = money(updated_user.get("wallet_balance"))

    purchase = {
        "purchase_id": reference,
        "user_id": user["_id"],
        "package_id": package["package_id"],
        "package_title": package.get("title", ""),
        "price": price,
        "currency": "GHS",
        "total_contacts": int(package.get("total_contacts", 0) or 0),
        "wallet_before": wallet_before,
        "wallet_after": wallet_after,
        "status": "completed",
        "created_at": now,
    }

    try:
        purchases_collection().insert_one(purchase)
    except DuplicateKeyError:
        users_collection().update_one({"_id": user["_id"]}, {"$inc": {"wallet_balance": price}, "$set": {"updated_at": now_utc()}})
        existing = purchases_collection().find_one({"user_id": user["_id"], "package_id": package["package_id"], "status": "completed"})
        return {"success": False, "message": "You have already purchased this contact package.", "purchase": safe_purchase(existing or purchase)}, 409

    wallet_transactions_collection().insert_one({
        "user_id": user["_id"],
        "type": "debit",
        "category": "contact_purchase",
        "amount": price,
        "currency": "GHS",
        "status": "success",
        "description": f"Purchased contact package: {package.get('title', '')}",
        "reason": f"Purchased contact package: {package.get('title', '')}",
        "reference": reference,
        "balance_before": wallet_before,
        "balance_after": wallet_after,
        "created_at": now,
        "updated_at": now,
    })

    source_contacts = list(marketplace_contacts_collection().find({"package_id": package["package_id"]}))
    existing_numbers = set()
    for contact in contacts_collection().find({"user_id": user["_id"]}, {"phone": 1, "normalized_phone": 1}):
        normalized = normalize_phone(contact.get("normalized_phone") or contact.get("phone", ""))
        if normalized:
            existing_numbers.add(normalized)

    imported = 0
    duplicate_skipped = 0
    failed = 0
    group_name = package.get("title", "")
    docs = []
    for item in source_contacts:
        normalized = normalize_phone(item.get("normalized_phone") or item.get("phone", ""))
        if not normalized:
            failed += 1
            continue
        if normalized in existing_numbers:
            duplicate_skipped += 1
            continue
        existing_numbers.add(normalized)
        docs.append({
            "user_id": user["_id"],
            "source": "marketplace",
            "source_package_id": package["package_id"],
            "group_name": group_name,
            "group": group_name,
            "name": "VireSender_purchase",
            "sender_id": "VireSender_purchase",
            "contact_name": item.get("name", ""),
            "phone": normalized,
            "normalized_phone": normalized,
            "email": item.get("email", ""),
            "age": "",
            "location": item.get("location", ""),
            "notes": item.get("notes", ""),
            "created_at": now,
            "updated_at": now,
        })

    if docs:
        try:
            result = contacts_collection().insert_many(docs, ordered=False)
            imported = len(result.inserted_ids)
        except Exception:
            for doc in docs:
                try:
                    contacts_collection().insert_one(doc)
                    imported += 1
                except DuplicateKeyError:
                    duplicate_skipped += 1
                except Exception:
                    failed += 1

    summary = {
        "total_contacts": len(source_contacts),
        "imported_contacts": imported,
        "duplicate_skipped": duplicate_skipped,
        "failed_contacts": failed,
    }
    purchases_collection().update_one({"purchase_id": reference}, {"$set": {"import_summary": summary}})
    purchase["import_summary"] = summary
    create_notification(
        user["_id"], "contacts", "Contact group purchased",
        f"Purchased {package.get('title', '')}. Imported {imported} contact(s).",
        "success", "contact_marketplace", package["package_id"], "/user/contacts",
        {"price": price, **summary},
    )
    return jsonify({
        "success": True,
        "message": f"Purchased {package.get('title', '')}. Imported {imported} contacts.",
        "purchase": safe_purchase(purchase),
        "import_summary": summary,
        "wallet_balance": wallet_after,
    })
