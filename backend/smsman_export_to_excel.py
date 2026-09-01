"""
SMS-MAN Export to Excel

Fetches SMS-MAN countries and services using your API token, then creates one Excel file
with separate tabs:
  1. Countries
  2. Services
  3. Raw Countries JSON
  4. Raw Services JSON

Setup:
  pip install requests openpyxl python-dotenv

.env:
  SMSMAN_API_TOKEN=your_api_token_here

Run:
  python smsman_export_to_excel.py
"""

import json
import os
from datetime import datetime
from typing import Any, Dict, List

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://api.sms-man.com/control"
OUTPUT_FILE = "smsman_countries_services.xlsx"


def get_token() -> str:
    load_dotenv()
    token = os.getenv("SMSMAN_API_TOKEN", "EMDQE326zbZW7c0MD3DW2weeruu8h").strip()
    if not token:
        raise RuntimeError("Missing SMSMAN_API_TOKEN in .env file")
    return token


def fetch_endpoint(endpoint: str, token: str) -> Any:
    url = f"{BASE_URL}/{endpoint}"
    response = requests.get(url, params={"token": token}, timeout=30)
    response.raise_for_status()

    try:
        data = response.json()
    except ValueError as exc:
        raise RuntimeError(f"SMS-MAN returned non-JSON response for {endpoint}: {response.text[:300]}") from exc

    if isinstance(data, dict) and data.get("success") is False:
        raise RuntimeError(f"SMS-MAN error for {endpoint}: {data}")

    return data


def normalize_rows(data: Any, preferred_columns: List[str]) -> List[Dict[str, Any]]:
    if isinstance(data, dict):
        data = list(data.values())

    rows: List[Dict[str, Any]] = []
    if not isinstance(data, list):
        return rows

    for item in data:
        if isinstance(item, dict):
            row = {col: item.get(col, "") for col in preferred_columns}
            for key, value in item.items():
                if key not in row:
                    row[key] = value
            rows.append(row)
        else:
            rows.append({"value": item})

    return rows


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 3, 12), 42)


def write_table(ws, rows: List[Dict[str, Any]]):
    if not rows:
        ws.append(["No data returned"])
        return

    headers = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    style_sheet(ws)


def write_raw_json(ws, title: str, data: Any):
    ws.append([title])
    ws.append([json.dumps(data, indent=2, ensure_ascii=False)])
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E40AF")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[2].height = 420
    ws.sheet_view.showGridLines = False


def create_excel(countries_data: Any, services_data: Any, output_file: str):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    countries_rows = normalize_rows(countries_data, ["id", "title"])
    services_rows = normalize_rows(services_data, ["id", "name", "code"])

    ws_countries = wb.create_sheet("Countries")
    write_table(ws_countries, countries_rows)

    ws_services = wb.create_sheet("Services")
    write_table(ws_services, services_rows)

    ws_raw_countries = wb.create_sheet("Raw Countries JSON")
    write_raw_json(ws_raw_countries, "Countries raw API response", countries_data)

    ws_raw_services = wb.create_sheet("Raw Services JSON")
    write_raw_json(ws_raw_services, "Services raw API response", services_data)

    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["SMS-MAN Export Summary", ""])
    ws_summary.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Countries Count", len(countries_rows)])
    ws_summary.append(["Services Count", len(services_rows)])
    ws_summary.append(["Countries Endpoint", f"{BASE_URL}/countries"])
    ws_summary.append(["Services Endpoint", f"{BASE_URL}/applications"])
    ws_summary["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws_summary["A1"].fill = PatternFill("solid", fgColor="1E40AF")
    ws_summary.merge_cells("A1:B1")
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 60
    ws_summary.sheet_view.showGridLines = False

    wb.save(output_file)


def main():
    token = get_token()
    print("Fetching countries...")
    countries = fetch_endpoint("countries", token)
    print("Fetching services...")
    services = fetch_endpoint("applications", token)

    create_excel(countries, services, OUTPUT_FILE)
    print(f"Done: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
